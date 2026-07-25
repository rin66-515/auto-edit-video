from io import BytesIO

from openpyxl import Workbook,load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment,Border,Font,PatternFill,Protection,Side
from openpyxl.worksheet.datavalidation import DataValidation


SHEET_NAME="字幕审核"
HEADERS=("序号","成片文件","开始时间","结束时间","中文字幕","日文字幕","处理方式","需复核","原识别","来源素材","校验ID")
EDITABLE_COLUMNS={3,4,5,6,7,8}
HEADER_FILL=PatternFill("solid",fgColor="245C8A")
HEADER_FONT=Font(color="FFFFFF",bold=True)
EDIT_FILL=PatternFill("solid",fgColor="EAF6EE")
READONLY_FILL=PatternFill("solid",fgColor="F2F4F7")
WARNING_FILL=PatternFill("solid",fgColor="FFF2CC")
THIN_GRAY=Side(style="thin",color="D8DEE8")


def build_caption_workbook(rows):
    workbook=Workbook();sheet=workbook.active;sheet.title=SHEET_NAME
    sheet.sheet_view.showGridLines=False;sheet.freeze_panes="A2"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append([row.get(header,"") for header in HEADERS])
    last=max(sheet.max_row,2)
    sheet.auto_filter.ref=f"A1:J{last}"
    widths={"A":8,"B":15,"C":17,"D":17,"E":34,"F":34,"G":13,"H":12,"I":30,"J":28,"K":26}
    for column,width in widths.items():sheet.column_dimensions[column].width=width
    sheet.column_dimensions["K"].hidden=True
    for cell in sheet[1]:
        cell.fill=HEADER_FILL;cell.font=HEADER_FONT;cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    sheet.row_dimensions[1].height=28
    for row_index in range(2,last+1):
        sheet.row_dimensions[row_index].height=42
        for column_index in range(1,12):
            cell=sheet.cell(row_index,column_index)
            cell.fill=EDIT_FILL if column_index in EDITABLE_COLUMNS else READONLY_FILL
            cell.alignment=Alignment(horizontal="center" if column_index in {1,2,3,4,7,8} else "left",vertical="center",wrap_text=True)
            cell.border=Border(bottom=THIN_GRAY)
            cell.protection=Protection(locked=column_index not in EDITABLE_COLUMNS)
    action_validation=DataValidation(type="list",formula1='"保留,省略"',allow_blank=False,error="只能选择“保留”或“省略”",errorTitle="处理方式无效")
    review_validation=DataValidation(type="list",formula1='"是,否"',allow_blank=False,error="只能选择“是”或“否”",errorTitle="复核状态无效")
    sheet.add_data_validation(action_validation);sheet.add_data_validation(review_validation)
    action_validation.add(f"G2:G{last}");review_validation.add(f"H2:H{last}")
    sheet.conditional_formatting.add(f"A2:J{last}",FormulaRule(formula=['$H2="是"'],fill=WARNING_FILL))
    notes=workbook.create_sheet("使用说明")
    notes.sheet_view.showGridLines=False;notes.column_dimensions["A"].width=110
    instructions=[
        "成片字幕人工审核说明",
        "1. 只修改绿色列：开始时间、结束时间、中文字幕、日文字幕、处理方式、需复核。",
        "2. 时间格式固定为 HH:MM:SS.mmm，例如 00:01:23.450。",
        "3. 不要删除行、插入行或修改隐藏的校验ID列；导入时会严格校验。",
        "4. 不需要的语气词或整句字幕，把“处理方式”改为“省略”。",
        "5. 校对完成后保存为 .xlsx，再回到审核页面导入同一版本。",
    ]
    for index,text in enumerate(instructions,1):
        cell=notes.cell(index,1,text);cell.alignment=Alignment(wrap_text=True,vertical="center")
        if index==1:cell.font=Font(size=16,bold=True,color="245C8A");notes.row_dimensions[index].height=30
        else:notes.row_dimensions[index].height=27
    output=BytesIO();workbook.save(output);return output.getvalue()


def read_caption_workbook(content):
    workbook=load_workbook(BytesIO(content),read_only=False,data_only=True)
    if SHEET_NAME not in workbook.sheetnames:raise ValueError(f"缺少工作表“{SHEET_NAME}”")
    sheet=workbook[SHEET_NAME]
    headers=tuple(str(sheet.cell(1,index).value or "").strip() for index in range(1,len(HEADERS)+1))
    if headers!=HEADERS:raise ValueError("字幕审核表头已被修改，请重新导出")
    rows=[]
    for row_index in range(2,sheet.max_row+1):
        values={header:sheet.cell(row_index,column_index).value for column_index,header in enumerate(HEADERS,1)}
        if not any(value not in {None,""} for value in values.values()):continue
        values["_excel_row"]=row_index;rows.append(values)
    return rows
