import base64
import json
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from fastapi import HTTPException
from openpyxl import load_workbook
from app import db
from app.main import CaptionWorkbookImportIn,export_caption_summary,export_version_caption_xlsx,import_version_caption_workbook
from app.media import _write_bilingual_srt
from _support import IsolatedDbTestCase


class CaptionWorkbookWorkflowTest(IsolatedDbTestCase):
    def setUp(self):
        super().setUp();stamp=db.now();self.slug=f"workbook-test-{id(self)}"
        self.project_id=db.execute("INSERT INTO projects(slug,title,source_dir,status,mode,created_at,updated_at) VALUES(?,?,?,?,?,?,?)",(self.slug,"Workbook Test",f"/vlog/inbox/{self.slug}","visual_failed","existing",stamp,stamp))
        db.create_control(self.project_id,"stopped","ready_for_visual","字幕人工复核")
        analysis={"caption_version":3,"bilingual_captions":[{"start":1.25,"end":2.75,"source":"こんにちは","zh":"你好","ja":"こんにちは","confidence":0.7,"needs_review":True}]}
        self.asset_id=db.execute("INSERT INTO assets(project_id,path,filename,bytes,analysis,created_at) VALUES(?,?,?,?,?,?)",(self.project_id,f"/vlog/inbox/{self.slug}/a.mp4","a.mp4",100,json.dumps(analysis,ensure_ascii=False),stamp))
        snapshot={"long":[{"asset_id":self.asset_id,"start":0.5,"end":3.5,"audio_mode":"dialogue"}]}
        self.export_id=db.execute("INSERT INTO exports(project_id,version,format,path,status,timeline_snapshot,created_at) VALUES(?,?,?,?,?,?,?)",(self.project_id,"v5","long_16x9",json.dumps([f"/vlog/_automation/outputs/{self.slug}/v5/long.mp4"]),"review_ready",json.dumps(snapshot,ensure_ascii=False),stamp))

    def test_final_cut_workbook_creates_new_non_destructive_render_version(self):
        response=export_version_caption_xlsx(self.export_id);workbook=load_workbook(BytesIO(response.body));sheet=workbook["字幕审核"]
        self.assertEqual(2,sheet.max_row);self.assertEqual("long",sheet["B2"].value);self.assertEqual("00:00:00.750",sheet["C2"].value);self.assertEqual("00:00:02.250",sheet["D2"].value);self.assertTrue(sheet.column_dimensions["K"].hidden)
        sheet["E2"]="你好，最终版";sheet["F2"]="こんにちは、完成版";sheet["H2"]="否"
        output=BytesIO();workbook.save(output)
        result=import_version_caption_workbook(self.export_id,CaptionWorkbookImportIn(filename="v5-edited.xlsx",xlsx_base64=base64.b64encode(output.getvalue()).decode()))
        self.assertEqual(1,result["changed_rows"]);self.assertEqual("v6",result["version"])
        original=db.row("SELECT status,caption_overrides FROM exports WHERE id=?",(self.export_id,));created=db.row("SELECT status,source_export_id,caption_overrides,timeline_snapshot FROM exports WHERE id=?",(result["new_export_id"],))
        self.assertEqual("review_ready",original["status"]);self.assertEqual("{}",original["caption_overrides"]);self.assertEqual("render_requested",created["status"]);self.assertEqual(self.export_id,created["source_export_id"]);self.assertIn("你好，最终版",created["caption_overrides"]);self.assertTrue(created["timeline_snapshot"])
        control=db.control(self.project_id);self.assertEqual("stopped",control["desired_state"]);self.assertEqual("render_requested",control["resume_status"])

    def test_final_cut_summary_uses_output_timeline_not_raw_asset_totals(self):
        summary=export_caption_summary(self.export_id)
        self.assertEqual("v5",summary["version"]);self.assertEqual(1,summary["total"]);self.assertEqual(1,summary["needs_review"])
        self.assertEqual([{"name":"long","captions":1}],summary["outputs"])
        self.assertEqual("00:00:00.750",summary["rows"][0]["start"]);self.assertEqual("a.mp4",summary["rows"][0]["asset"])

    def test_workbook_rejects_time_beyond_output_without_partial_update(self):
        response=export_version_caption_xlsx(self.export_id);workbook=load_workbook(BytesIO(response.body));workbook["字幕审核"]["D2"]="00:00:09.000"
        output=BytesIO();workbook.save(output)
        body=CaptionWorkbookImportIn(filename="bad-time.xlsx",xlsx_base64=base64.b64encode(output.getvalue()).decode())
        with self.assertRaises(HTTPException):import_version_caption_workbook(self.export_id,body)
        self.assertEqual("{}",db.row("SELECT caption_overrides FROM exports WHERE id=?",(self.export_id,))["caption_overrides"])

    def test_unchanged_existing_sub_tenth_caption_can_be_imported(self):
        analysis={"caption_version":3,"bilingual_captions":[{"start":0.6,"end":0.68,"source":"没","zh":"没","ja":"ない"}]}
        db.execute("UPDATE assets SET analysis=? WHERE id=?",(json.dumps(analysis,ensure_ascii=False),self.asset_id))
        response=export_version_caption_xlsx(self.export_id);workbook=load_workbook(BytesIO(response.body));output=BytesIO();workbook.save(output)
        result=import_version_caption_workbook(self.export_id,CaptionWorkbookImportIn(filename="unchanged.xlsx",xlsx_base64=base64.b64encode(output.getvalue()).decode()))
        self.assertEqual(0,result["changed_rows"]);self.assertEqual("XLSX与该成片版本字幕一致，没有需要保存的修改",result["message"])

    def test_final_cut_workbook_merges_split_caption_and_allows_timing_edit(self):
        analysis={"caption_version":3,"bilingual_captions":[
            {"start":0.6,"end":0.9,"source":"うん","zh":"嗯","ja":"うん"},
            {"start":1.25,"end":2.75,"source":"こんにちは","zh":"你好","ja":"こんにちは","needs_review":True},
        ]}
        db.execute("UPDATE assets SET analysis=? WHERE id=?",(json.dumps(analysis,ensure_ascii=False),self.asset_id))
        snapshot={"long":[
            {"asset_id":self.asset_id,"start":0.5,"end":2.0,"audio_mode":"dialogue"},
            {"asset_id":self.asset_id,"start":2.0,"end":3.5,"audio_mode":"dialogue"},
        ]}
        db.execute("UPDATE exports SET timeline_snapshot=? WHERE id=?",(json.dumps(snapshot,ensure_ascii=False),self.export_id))
        response=export_version_caption_xlsx(self.export_id);workbook=load_workbook(BytesIO(response.body));sheet=workbook["字幕审核"]
        self.assertEqual(2,sheet.max_row);self.assertEqual("00:00:00.750",sheet["C2"].value);self.assertEqual("00:00:02.250",sheet["D2"].value)
        sheet["C2"]="00:00:00.800";sheet["D2"]="00:00:02.100";sheet["E2"]="合并后的字幕";sheet["F2"]="結合した字幕";sheet["H2"]="否"
        output=BytesIO();workbook.save(output)
        result=import_version_caption_workbook(self.export_id,CaptionWorkbookImportIn(filename="merged.xlsx",xlsx_base64=base64.b64encode(output.getvalue()).decode()))
        self.assertEqual(1,result["changed_rows"]);self.assertEqual(2,result["changed_segments"])
        overrides=json.loads(db.row("SELECT caption_overrides FROM exports WHERE id=?",(result["new_export_id"],))["caption_overrides"])
        self.assertEqual(2,len(overrides))
        timing=[value for value in overrides.values() if value.get("timing_adjusted")][0];continuation=[value for value in overrides.values() if value.get("timing_continuation_omitted")][0]
        self.assertEqual(0.8,timing["output_start"]);self.assertEqual(2.1,timing["output_end"]);self.assertEqual("合并后的字幕",timing["zh"]);self.assertTrue(continuation["omit"])

    def test_rendered_srt_uses_same_merge_and_filler_cleanup(self):
        analysis={"bilingual_captions":[
            {"start":0.6,"end":0.9,"zh":"嗯","ja":"うん"},
            {"start":1.25,"end":2.75,"zh":"你好","ja":"こんにちは"},
        ]}
        asset={"id":self.asset_id,"analysis":json.dumps(analysis,ensure_ascii=False)}
        used=[
            ({"start":0.5},asset,1.5,1),
            ({"start":2.0},asset,1.5,2),
        ]
        overrides={
            f"long|1|{self.asset_id}|2":{"zh":"调整时间","ja":"時間調整","output_start":0.8,"output_end":2.1},
            f"long|2|{self.asset_id}|2":{"omit":True},
        }
        with tempfile.TemporaryDirectory() as directory:
            target=Path(directory)/"captions.srt";self.assertTrue(_write_bilingual_srt(used,target,"long_16x9",overrides,"long"))
            text=target.read_text(encoding="utf-8-sig")
        self.assertEqual(1,text.count("调整时间"));self.assertNotIn("嗯",text);self.assertIn("00:00:00,800 --> 00:00:02,100",text)


if __name__=="__main__":unittest.main()
